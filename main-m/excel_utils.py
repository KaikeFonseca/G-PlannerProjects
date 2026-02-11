# excel_utils.py

import os
import win32com.client
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def updateExcel(caminho_arquivo):
    """Atualiza todas as conexões de dados de um arquivo Excel."""
    if not os.path.exists(caminho_arquivo):
        print(f"[ERRO] O arquivo Excel não foi encontrado em: {caminho_arquivo}")
        return False

    print("Iniciando a atualização da planilha Excel...")
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        workbook = excel.Workbooks.Open(caminho_arquivo)
        
        print("Configurando conexões para atualização síncrona...")
        for conn in workbook.Connections:
            conn.OLEDBConnection.BackgroundQuery = False
            
        print("Atualizando conexões de dados... Isso pode levar alguns minutos.")
        workbook.RefreshAll()
        excel.Calculate()
        
        print("Salvando o arquivo...")
        workbook.Save()
        workbook.Close(SaveChanges=True)
        print("Planilha atualizada com sucesso!")
        return True
    except Exception as e:
        print(f"[ERRO] Falha ao atualizar a planilha Excel: {e}")
        return False
    finally:
        if excel:
            excel.Quit()

def formatar_como_tabela(df, caminho_arquivo, nome_planilha="Sheet1", nome_tabela="Tabela"):
    """Exporta um DataFrame para Excel e aplica formatação como Tabela."""
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=nome_planilha, index=False)

    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_planilha]
    
    max_row = ws.max_row
    max_col = ws.max_column
    intervalo = f"A1:{get_column_letter(max_col)}{max_row}"
    
    tabela = Table(displayName=nome_tabela, ref=intervalo)
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    
    wb.save(caminho_arquivo)
    print(f"Tabela '{nome_tabela}' salva com sucesso em '{caminho_arquivo}'!")