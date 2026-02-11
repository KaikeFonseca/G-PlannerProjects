import os
import win32com.client
import time
import datetime
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

#IMPORTADO PELO RNC

from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext


#OUTRO ARQUIVO
from mb52 import updateStock
#IDEM
from montar_patan_logic import montar_patan

#CAMINHOS
PATAN = r"\\sb2-fs\11_GESTAO_DA_LOGISTICA$\LOGISTICA\104 - AutomacaoPlanner\Automação Planner\AutomacaoPlanner.xlsx"
EXCEL = r"\\sb2-fs\4_GESTAO_DA_QUALIDADE$\00_KPI_&_GSD\Melhoria Contínua\Fluxos\014. Planner Project\excels"

# Variável global para armazenar o nome do último diário de bordo gerado
last_generated_output_file = None # Changed to store the output_patan file

# CALCULAR DIFERENÇA DE HORAS
def diferença_horas(h1, h2):
    # Define o formato
    formato = "%H:%M:%S"
    
    # Converte as strings para datetime
    inicio = datetime.strptime(h1, formato)
    fim = datetime.strptime(h2, formato)
    
    # Calcula a diferença
    if fim < inicio:
        # Se a hora de fim é menor que a de início, adicionamos um dia
        fim += timedelta(days=1)
    
    diferença = fim - inicio
    
    # Retorna a diferença no formato H:M:S
    horas_total, resto = divmod(diferença.total_seconds(), 3600)
    minutos, segundos = divmod(resto, 60)
    
    return f"{int(horas_total):02}:{int(minutos):02}:{int(segundos):02}"
#--------------------------------------------------------------------------------------------------#

#CONFIGURAR SHAREPOINT - MAIN
#------------------------------------#
def configuracoes_sharepoint():
    url='https://gitservices.sharepoint.com/sites/LogsticaAA3/'
    client_id ='sb2mcontinua@br.gestamp.com'
    client_secret = '@Gestamp08'

    app_settings = {
        'url': url,
        'client_id': client_id,
        'client_secret': client_secret
    }
    return url, client_id, client_secret, app_settings
#--------------------------------------------------------------------------------------------------#

#FUNÇÕES DO DATAFRAME
#------------------------------------#
##DF - FLUXO PRINCIPAL - COM FILTRO E RETORNO DE BASE
def dataframe_lista_rastreabilidade_INICIAL():
    url, client_id, client_secret, app_settings = configuracoes_sharepoint()
    # Configuração do contexto de autenticação
    ctx_auth = AuthenticationContext(url)
    if ctx_auth.acquire_token_for_user(app_settings['client_id'], app_settings['client_secret']):
        ctx = ClientContext(app_settings['url'], ctx_auth)
        list_obj = ctx.web.lists.get_by_title('Lista de Rastreabilidade') # Acessando a lista desejada
        #item_properties = {'Title': 'Novo Item 2','true_false': False} # Adicionando um novo item à lista # Adicione mais propriedades de item conforme necessário
        #list_obj.add_item(item_properties)
        ctx.load(list_obj)
        items = list_obj.get_items().top(1000000).get().execute_query()

        #for item in items:
        #    pprint(item.properties)
        # Selecionar apenas as colunas desejadas
        colunas_desejadas = ['Title', 'ID', 'id_rnc', 'comando', 'id_comando', 'doc_rastreabilidade', 
                            'cod_rastreabilidade', 'cod_produto', 'cod_comp', 
                            'data_reclamacao', 'data_inicial', 'data_final', 'ocorrencia', 'seq_relacionada', 'defeito', 'final_comp']
        
        # Filtrar os itens com status 'Z' OR 'G' e criar DataFrame
        data = []
        for item in items:
            if item.properties['status_geral'] == 'Z' or item.properties['Title'] == 'GG' or item.properties['Title'] == 'Z':
                item_data = {col: item.properties.get(col, None) for col in colunas_desejadas}
                data.append(item_data)

        # Criar o DataFrame
        df = pd.DataFrame(data)

        # Renomear a coluna 'Title' para 'status'
        df.rename(columns={'Title': 'status'}, inplace=True)

        #print(df)
        # Exportar o DataFrame para Excel na pasta de downloads do usuário
        #caminho_download = str(Path.home() / "Downloads" / "dados_sharepoint.xlsx")
        #df.to_excel(caminho_download, index=False)
        #print(f"DataFrame exportado para: {caminho_download}")
        return df, ctx, list_obj
    else:
        print(ctx_auth.get_last_error())
        print('Erro na autenticação')

##DF - EDITA A LINHA CASO ELA JA EXISTA (CAUSA: SEERROR)
def dataframe_editar_item_sharepoint(items, ID, novos_valores, ctx2):
    try:
        # Filtrar os itens que correspondem ao ID informado
        items_to_edit = []
        for item in items:
            if item.properties.get('ID') == int(ID):
                items_to_edit.append(item)

        #
        item_to_edit = items_to_edit[0]
        
        # Atualizar os campos do item com os novos valores
        for key, value in novos_valores.items():
            item_to_edit.set_property(key, value)
        
        item_to_edit.update()  # Marcar o item para atualização
        ctx2.execute_query()  # Confirma a atualização no SharePoint
            
        print(f"\n\t\t---------------------------------------")
        print(f"\t\tItem com ID={ID} foi editado com sucesso.")
        print(f"\t\t---------------------------------------\n")
        return True
    except Exception as e:
        print(f"\n\t\t---------------------------------------")
        print(f"Erro ao editar item no sharepoint. ID: {ID}\nError: {e}")
        print(f"\t\t---------------------------------------\n")
        return False

# Função para iterar sobre o DataFrame e enviar para o SharePoint
def atualizar_df_sharepoint(df_grouped, items, ctx2):
    try:
        for index, row in df_grouped.iterrows():
            # Pegando o ID do item para o update
            item_id = row['ID']
            
            # Cria um dicionário com os campos que precisam ser atualizados
            novos_valores = {
                'idrnc': row['idrnc'],
                'Title': row['Sequencia Relacionada'],
                'etiquetasap': row['camposJuntos'],
                # Outros campos que você deseja atualizar
            }
            
            # Chama a função para editar o item no SharePoint
            auxBool = dataframe_editar_item_sharepoint(items, item_id, novos_valores, ctx2)
            return auxBool
    except Exception as e:
        print(f"\nNão foi possivel atualizar o sharepoint - ETIQUETAS VIVAS - VERIFICAR. \nError: {e}")
        return False
#--------------------------------------------------------------------------------------------------#


def clear_screen():
    """Limpa a tela do terminal."""
    os.system("cls" if os.name == "nt" else "clear")

def main_menu():
    """Exibe o menu principal e retorna a escolha do usuário."""
    clear_screen()
    print("\nMenu Principal:")
    print("1 - Montar Patan")
    print("2 - Loopar")
    print("0 - Sair")

    choice = input("\nEscolha uma opção: ")
    return choice

def montar_patan_menu():
    """Gerencia o processo de montagem de Patan, incluindo sub-menus e geração de diário de bordo."""
    global last_generated_output_file
    clear_screen()
    print("\n--- Montar Patan ---")
    while True:
        letra_patan = input("Informe a letra patan (A até D): ").upper()
        if letra_patan in ["A", "B", "C", "D"]:
            break
        else:
            print("Letra Patan inválida. Por favor, digite A, B, C ou D.")
    
    clear_screen()
    while True:
        print("\nSelecione a linha:")
        print("1 - LINHA 1")
        print("2 - LINHA 2")
        print("3 - LINHA 3")
        linha = input("Escolha uma opção: ")
        if linha in ["1", "2", "3"]:
            break
        else:
            print("Opção de linha inválida. Por favor, digite 1, 2 ou 3.")

    clear_screen()
    while True:
        print("\nInforme o Turno:")
        print("1 - Turno 1")
        print("2 - Turno 2")
        print("3 - Turno 3")
        turno = input("Escolha uma opção: ")
        if turno in ["1", "2", "3"]:
            break
        else:
            print("Opção de turno inválida. Por favor, digite 1, 2 ou 3.")

    clear_screen()
    print("ATUALIZANDO O ESTOQUE - AGUARDE UM MOMENTO!")
    updateStock()
    clear_screen()
    print("ATUALIZANDO O ARQUIVO PRINCIPAL")
    updateExcel(PATAN)
    print(f"\nProcessando para Patan {letra_patan}, Linha {linha}, Turno {turno}...")
    
    df_result, df_diario_de_bordo, df_com_erros = montar_patan(letra_patan, linha, turno, PATAN)

    print("\n--- DataFrame de Saída ---")
    print(df_result.to_string())

    print("\n--- Diário de Bordo ---")
    print(df_diario_de_bordo.to_string())

    print("\n--- Erros Encontrados ---")
    print(df_com_erros.to_string())

    # Salvar resultados em arquivos para verificação
    output_filename = EXCEL + f'/TEMPORARIO---PATAN-{letra_patan}-LINHA-{linha}-TURNO{turno}.xlsx'
    diary_filename = EXCEL + f'/diarioBordo/diario_de_bordo_{letra_patan}_Linha{linha}_Turno{turno}.xlsx'
    errors_filename = EXCEL + f'/erros/erros_processamento_{letra_patan}_Linha{linha}_Turno{turno}.xlsx'

    df_result.to_excel(output_filename, index=False)
    df_diario_de_bordo.to_excel(diary_filename, index=False)
    df_com_erros.to_excel(errors_filename, index=False)
    
    print(f"\nResultados salvos em {output_filename}, {diary_filename} e {errors_filename}")
    last_generated_output_file = output_filename # Atualiza a variável global com o nome do arquivo de saída

    clear_screen()
    while True:
        auxAnswer = input("Deseja abrir o Patan Temporario? (S/N)").upper()
        if auxAnswer in ['S', 'N']:
            if auxAnswer == 'N':
                break
            else:
                os.startfile(output_filename)
                break
        else:
            print("Resposta inválida. Por favor, digite S para sim e N para não.")


    clear_screen()
    input("Feche o Excel para continuar. Pressione enter assim que fechar.")
    df_ExcelEditado = pd.read_excel(output_filename)
    createWoorkSheetPlanner(df_ExcelEditado,linha)

def loop():
    """Visualiza o conteúdo do arquivo Excel de estoque."""
    clear_screen()
    print("\n--- Visualizar Estoque ---")
    # Caminho fictício, pois o arquivo não foi fornecido
    excel_file = "/home/ubuntu/estoque-teste-EXCEL.xlsx" 
    if os.path.exists(excel_file):
        try:
            df = pd.read_excel(excel_file)
            print("\nConteúdo do arquivo Estoque-Teste-EXCEl.XLSX:")
            print(df.to_string())
        except Exception as e:
            print(f"Erro ao ler o arquivo Excel: {e}")
    else:
        print(f"Arquivo \'{excel_file}\' não encontrado. Certifique-se de que o arquivo está no mesmo diretório do script ou forneça o caminho correto.")
    input("Pressione Enter para continuar...")

def createWoorkSheetPlanner(df_output, linha:str):
    df_worksheet_planner = pd.DataFrame(columns=[
        "Material", "posto", "tags", "descricao", "checklist", "linha", "data1", "data2"
    ])

    today = datetime.today().date()
    data1_str = today.strftime("%d/%m")
    data2_str = today.strftime("%Y-%d-%m")
    checklist_fixed = "1 - (ABASTECIMENTO) ESTAMPADO ABASTECIDO;2 - (ABASTECIMENTO) EMBALAGEM E VTOV  ABASTECIDOS;3 - (PRODUÇÃO) PRÉ-SETUP;4 - (PRODUÇÃO) FECHAMENTO;5 - LIMPEZA WIP2 OU DTR3"

    # Agrupar por posto para inserir a linha extra
    for posto_name, group in df_output.groupby("posto"):
        # Inserir linha extra antes da primeira linha de sequência 1
        first_row_of_group = group.iloc[0]
        turno_extra = first_row_of_group["turno"]
        posto_extra = first_row_of_group["posto"]
        posto_num = posto_extra[-2:] # Últimos 2 dígitos do posto

        df_worksheet_planner.loc[len(df_worksheet_planner)] = {
            "Material": f"{turno_extra}° TURNO - {data1_str}",
            "posto": posto_extra,
            "tags": f"PONT. {posto_num}",
            "descricao": None,
            "checklist": None,
            "linha": int(linha)
            #"data1": data1_str,
            #"data2": None # User specified only data1 for this extra row
        }

        for idx, row in group.iterrows():
            material = row["Material"]
            posto = row["posto"]
            patan = row["patan"]
            turno = row["turno"]
            prod_em_linha = row["prodEmLinha"]
            hora_prod_inicial = row["horaProdInicial"]
            hora_prod_final = row["horaProdFinal"]
            tempo_prod = int(row["tempoProd"]) # Apenas a parte inteira
            kanbans = row["kanbans"]
            qtd_pecas_serem_produzidas = row["qtdPecasSeremProduzidas"]
            comp_comb = row["compComb"]

            # Construir \'tags\'
            tags_list = []
            tags_list.append(f"PATAN ({patan})")
            tags_list.append(f"{turno}° TURNO")
            posto_num = posto[-2:]
            tags_list.append(f"PONT. {posto_num}")
            if prod_em_linha == 1:
                tags_list.append("PROD. EM LINHA")
            tags = ";;;".join(tags_list)

            # Construir \'descricao\'
            descricao = ""
            if prod_em_linha == 0:
                descricao += f"INICIO: {hora_prod_inicial}\n"
                descricao += f"FIM: {hora_prod_final}\n"
                descricao += f"TEMPO/PRODUÇÃO: {tempo_prod} MIN.\n"
                descricao += f"QTD. - {kanbans} K = {qtd_pecas_serem_produzidas} PÇS\n"
                
                # Processar compComb para descricao
                if comp_comb and comp_comb != 'nan':
                    componentes = comp_comb.split('\n')
                    for comp_info in componentes:
                        comp_info = comp_info.strip()
                        if not comp_info: # Skip empty strings
                            continue
                        try:
                            comp_parts = comp_info.split(':')
                            comp_nome = comp_parts[0].strip()
                            comp_qtd = int(float(comp_parts[1].strip())) # Convert to float first, then int
                            
                            tipo_componente = ""
                            if 'E' in comp_nome.upper(): # Check if 'E' is in component name
                                tipo_componente = "(ESTAMPADO)"
                            else:
                                tipo_componente = "(VtoV)"
                            descricao += f"{tipo_componente}: {comp_nome} - {comp_qtd} PÇS\n"
                        except (ValueError, IndexError) as e:
                            descricao += f"Erro ao processar componente: {comp_info} ({e})\n"
            else: # prodEmLinha == 1
                descricao = "PRODUÇÃO EM LINHA\n."

            df_worksheet_planner.loc[len(df_worksheet_planner)] = {
                "Material": material,
                "posto": posto,
                "tags": tags,
                "descricao": descricao,
                "checklist": checklist_fixed,
                "linha": int(linha)
                #"data1": data1_str,
                #"data2": data2_str
            }
    # Ordenar pela coluna "Material" em ordem decrescente
    df_worksheet_planner = df_worksheet_planner.sort_values(by="Material", ascending=False).reset_index(drop=True)

    output_worksheet_filename = r'C:\Users\sb2mcontinua\Gestamp Servicios S.A\Melhoria Contínua AA3 - Documentos\\DISPARADOR - PLANNER\aux-planner.xlsx'
    formatar_como_tabela(df_worksheet_planner, output_worksheet_filename, nome_tabela="Sheet")
    input("Pressione Enter para continuar...")

def formatar_como_tabela(df, caminho_arquivo, nome_planilha="Sheet1", nome_tabela="Tabela"):
    """
    Exporta um DataFrame para Excel e aplica formatação como Tabela do Excel.
    
    Parâmetros:
    - df: DataFrame pandas a ser exportado.
    - caminho_arquivo: caminho do arquivo .xlsx a ser salvo.
    - nome_planilha: nome da aba onde será salva a tabela (padrão: Sheet1).
    - nome_tabela: nome da tabela no Excel (padrão: Tabela).
    """
    # Salvar o DataFrame no Excel
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=nome_planilha, index=False)

    # Reabrir com openpyxl para aplicar formatação de tabela
    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_planilha]

    # Definir intervalo da tabela
    max_row = ws.max_row
    max_col = ws.max_column
    ultima_coluna = get_column_letter(max_col)
    intervalo = f"A1:{ultima_coluna}{max_row}"

    # Criar a tabela
    tabela = Table(displayName=nome_tabela, ref=intervalo)
    estilo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabela.tableStyleInfo = estilo

    ws.add_table(tabela)
    wb.save(caminho_arquivo)
    print(f"Tabela '{nome_tabela}' salva com sucesso em '{caminho_arquivo}'!")


def updateExcel(caminho_arquivo):
    """
    Usa a automação COM do Windows para abrir o Excel, atualizar todas as conexões
    de dados de forma síncrona e salvar o arquivo.
    """
    # Verifica se o arquivo existe
    if not os.path.exists(caminho_arquivo):
        print(f"[ERRO] O arquivo Excel não foi encontrado em: {caminho_arquivo}")
        return False

    print("Iniciando a atualização da planilha Excel...")
    excel = None  # Inicializa a variável para o bloco finally
    try:
        # Inicia uma instância do Excel
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Não mostra a janela do Excel

        # Abre o workbook
        workbook = excel.Workbooks.Open(caminho_arquivo)

        # Desabilita a atualização em segundo plano para garantir que o script espere
        print("Configurando conexões para atualização síncrona...")
        for conn in workbook.Connections:
            conn.OLEDBConnection.BackgroundQuery = False

        # Atualiza todas as conexões de dados
        print("Atualizando conexões de dados... Isso pode levar alguns minutos.")
        workbook.RefreshAll()
        
        # O script irá pausar aqui até que workbook.RefreshAll() seja concluído,
        # pois BackgroundQuery está definido como False.

        print("Atualização das conexões de dados concluída.")

        # Recalcula a pasta de trabalho, caso haja fórmulas dependentes dos dados atualizados.
        # Embora o RefreshAll() geralmente acione o cálculo, é uma boa prática garantir.
        excel.Calculate()

        # Salva o arquivo
        print("Salvando o arquivo...")
        workbook.Save()

        # Fecha o workbook e o Excel
        workbook.Close(SaveChanges=True)
        
        print("Planilha atualizada com sucesso!")
        return True

    except Exception as e:
        print(f"[ERRO] Falha ao atualizar a planilha Excel: {e}")
        return False
        
    finally:
        # Garante que o processo do Excel seja fechado em qualquer cenário
        if excel:
            excel.Quit()

def run_script():
    """Função principal que executa o loop do menu."""
    while True:
        choice = main_menu()

        if choice == "1":
            montar_patan_menu()
        elif choice == "2":
            loop()
        elif choice == "5":
            if last_generated_output_file:
                df_output_from_patan = pd.read_excel(last_generated_output_file)
                createWoorkSheetPlanner(df_output_from_patan)

        elif choice == "0":
            print("Saindo...")
            break
        else:
            print("Opção inválida. Tente novamente.")
            input("Pressione Enter para continuar...")

if __name__ == "__main__":
    #run_script()
    updateExcel(PATAN)


